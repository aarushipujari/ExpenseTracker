import os
import csv
import io
import json
import re
import sqlite3
from datetime import datetime, timedelta, date
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash, Response, jsonify
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from werkzeug.security import generate_password_hash, check_password_hash


app = Flask(__name__)

app.secret_key = os.environ.get(
    "EXPENSE_TRACKER_SECRET_KEY",
    "expense_tracker_secret_key_2026"
)

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DATABASE = os.path.join(BASE_DIR, "expenses.db")


# ============================================================
# INDIAN CURRENCY VALUE SYSTEM & CELL FORMATTER
# ============================================================

def format_indian_currency(value):
    if value is None:
        return "0"
    try:
        num = round(float(value))
    except (ValueError, TypeError):
        return str(value)
    
    is_negative = num < 0
    num_str = str(abs(num))
    
    if len(num_str) <= 3:
        res = num_str
    else:
        last_three = num_str[-3:]
        remaining = num_str[:-3]
        
        groups = []
        while len(remaining) > 2:
            groups.append(remaining[-2:])
            remaining = remaining[:-2]
        if remaining:
            groups.append(remaining)
        
        groups.reverse()
        res = ",".join(groups) + "," + last_three
        
    return f"-{res}" if is_negative else res


def clean_cell_display(val, header=""):
    if val is None:
        return ""
    val_str = str(val).strip()
    if not val_str:
        return ""
    
    s = val_str.replace(",", "").replace("₹", "").replace("$", "").replace("%", "").strip()
    try:
        num = float(s)
    except ValueError:
        return val_str
    
    h_lower = str(header).lower().strip()
    
    is_units_or_nav = any(k in h_lower for k in ["unit balance", "units", "unit", "nav", "quantity", "qty", "shares", "ratio", "isin", "folio", "sn", "sr no", "no."])
    
    if is_units_or_nav:
        rounded = round(num, 4)
        if rounded.is_integer():
            return str(int(rounded))
        return f"{rounded:.4f}".rstrip("0").rstrip(".")
    
    is_amount_col = any(k in h_lower for k in ["amount", "value", "cost", "market", "p&l", "pnl", "gain", "loss", "invest", "balance", "price", "total"])
    
    if is_amount_col or abs(num) >= 100:
        return format_indian_currency(round(num))
    elif num.is_integer():
        return str(int(num))
    else:
        rounded = round(num, 2)
        return str(int(rounded)) if rounded.is_integer() else f"{rounded:.2f}"


app.jinja_env.filters['inr'] = format_indian_currency
app.jinja_env.filters['indian_format'] = format_indian_currency
app.jinja_env.filters['format_cell'] = clean_cell_display


# ============================================================
# DATABASE
# ============================================================

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()

    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            date TEXT NOT NULL,
            type TEXT NOT NULL,
            category TEXT NOT NULL,
            subcategory TEXT,
            amount REAL NOT NULL,
            description TEXT,
            place TEXT,
            account TEXT,
            remarks TEXT
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS holdings_statements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            filename TEXT NOT NULL,
            title TEXT NOT NULL,
            upload_date TEXT NOT NULL,
            statement_date TEXT,
            invested_value REAL DEFAULT 0,
            present_value REAL DEFAULT 0,
            unrealized_pnl REAL DEFAULT 0,
            pnl_percentage REAL DEFAULT 0,
            has_pnl INTEGER DEFAULT 1,
            headers_json TEXT NOT NULL,
            data_json TEXT NOT NULL,
            row_count INTEGER DEFAULT 0
        )
    """)

    columns = [col["name"] for col in conn.execute("PRAGMA table_info(holdings_statements)").fetchall()]
    if "has_pnl" not in columns:
        conn.execute("ALTER TABLE holdings_statements ADD COLUMN has_pnl INTEGER DEFAULT 1")

    tx_columns = [col["name"] for col in conn.execute("PRAGMA table_info(transactions)").fetchall()]
    for col in ["user_id", "subcategory", "place", "account", "remarks"]:
        if col not in tx_columns:
            conn.execute(f"ALTER TABLE transactions ADD COLUMN {col} TEXT" if col != "user_id" else "ALTER TABLE transactions ADD COLUMN user_id INTEGER")

    conn.commit()
    conn.close()


# Ensure database and tables are created upon loading (vital for WSGI/Production servers)
try:
    init_db()
except Exception as e:
    pass


# ============================================================
# LOGIN REQUIRED DECORATOR
# ============================================================

def login_required(function):
    @wraps(function)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            flash("Please log in to continue.", "info")
            return redirect(url_for("login"))
        return function(*args, **kwargs)
    return wrapper


# ============================================================
# HELPERS
# ============================================================

def clean_amount(value):
    try:
        if value is None:
            return 0.0
        val_str = str(value).replace(",", "").replace("₹", "").replace("$", "").replace("€", "").replace("£", "").strip()
        return round(float(val_str), 2)
    except (ValueError, TypeError):
        return 0.0


def clean_num(val):
    if val is None:
        return None
    s = str(val).replace(",", "").replace("₹", "").replace("$", "").replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def clean_type(value):
    val = str(value or "").strip().lower()
    if val in ["income", "credit", "salary", "dividend", "interest"]:
        return "Income"
    if val in ["investment", "investments", "sip", "mutual fund", "mutual funds", "stocks", "insurance", "bonds", "fd", "ppf"]:
        return "Investment"
    return "Expense"


def is_investment_category(category):
    cat = str(category or "").strip().lower()
    return cat in [
        "investment", "investments", "sip", "mutual fund", "mutual funds",
        "stocks", "equity", "bonds", "crypto", "gold", "insurance",
        "fixed deposit", "fd", "ppf", "nps", "reit"
    ]


def get_real_type(transaction_type, category):
    category_str = str(category or "").strip()
    if is_investment_category(category_str):
        return "Investment"
    return clean_type(transaction_type)


def normalize_date(value):
    if value is None or str(value).strip() == "":
        return datetime.now().strftime("%Y-%m-%d")

    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")

    val_str = str(value).strip()

    if " " in val_str:
        date_part = val_str.split(" ")[0]
        if len(date_part) == 10 and date_part.count("-") == 2:
            return date_part

    if len(val_str) == 10 and val_str[4] == "-" and val_str[7] == "-":
        return val_str

    try:
        num_val = float(val_str)
        if 20000 < num_val < 60000:
            converted = datetime(1899, 12, 30) + timedelta(days=num_val)
            return converted.strftime("%Y-%m-%d")
    except ValueError:
        pass

    formats = [
        "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
        "%d/%m/%y", "%d-%m-%y",
        "%Y/%m/%d", "%Y-%m-%d",
        "%m/%d/%Y", "%m-%d-%Y",
        "%d %b %Y", "%d %B %Y",
        "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"
    ]

    for fmt in formats:
        try:
            parsed = datetime.strptime(val_str, fmt)
            return parsed.strftime("%Y-%m-%d")
        except ValueError:
            continue

    return val_str


def get_investment_category_and_sub(row):
    cat = str(row['category'] or '').strip()
    subcat = str(row['subcategory'] or '').strip()
    remarks = str(row['remarks'] or '').strip()
    desc = str(row['description'] or '').strip()
    place = str(row['place'] or '').strip()
    
    lower_all = f"{cat} {subcat} {remarks} {desc} {place}".lower()
    
    if any(k in lower_all for k in ['insurance', 'hdfcstandard', 'hdfc life', 'lic', 'max life', 'term plan', 'policy', 'click 2 invest']):
        inv_category = 'Insurance'
    elif any(k in lower_all for k in ['mutual fund', 'mf sip', 'sip', 'flexi cap', 'small cap', 'multi-asset', 'hybrid', 'focused fund', 'miraeasset', 'parag parikh', 'sbi small cap', 'advisors', 'uti', 'icici prudential', 'hdfc focused', 'fund']):
        inv_category = 'Mutual Funds'
    elif any(k in lower_all for k in ['stock', 'equity', 'zerodha', 'groww', 'share', 'shares']):
        inv_category = 'Stocks / Equity'
    elif any(k in lower_all for k in ['fixed deposit', 'flexi deposit', 'fd ', 'term deposit']):
        inv_category = 'Fixed Deposit'
    elif any(k in lower_all for k in ['bond', 'bonds', 'debenture']):
        inv_category = 'Bonds'
    elif any(k in lower_all for k in ['ppf', 'nps', 'epf', 'provident fund']):
        inv_category = 'PPF / NPS'
    elif any(k in lower_all for k in ['gold', 'sovereign gold', 'sgb']):
        inv_category = 'Gold'
    elif any(k in lower_all for k in ['crypto', 'bitcoin', 'ethereum']):
        inv_category = 'Crypto'
    elif any(k in lower_all for k in ['real estate', 'reit', 'property']):
        inv_category = 'Real Estate'
    elif cat and cat.lower() not in ['investment', 'investments', 'expense']:
        inv_category = cat
    elif subcat and subcat.lower() not in ['investment', 'investments']:
        inv_category = subcat
    else:
        inv_category = 'Other Investments'

    if subcat and subcat != inv_category and subcat.lower() not in ['investment', 'investments', 'mutual funds', 'mf sip', 'sip']:
        scheme = subcat
    elif remarks and not remarks.lower().startswith("ach-dr-") and not remarks.lower().startswith("autobpay/"):
        scheme = remarks
    elif desc and not desc.startswith("ACH-DR-") and not desc.startswith("UPI/") and not desc.startswith("AUTOBPay/"):
        scheme = desc
    elif remarks:
        scheme = remarks
    else:
        scheme = inv_category
        
    return inv_category, scheme


# ============================================================
# SMART HOLDINGS & PORTFOLIO FILE PARSER
# ============================================================

def is_row_total_summary(row_cells):
    for cell in row_cells:
        c_str = str(cell).strip().lower()
        if c_str in ["total", "grand total", "summary", "total:", "subtotal", "sub total", "grand total:"]:
            return True
        if (c_str.startswith("total") or c_str.startswith("grand total")) and len(c_str) <= 20:
            return True
    return False


def parse_holdings_data(raw_rows, filename):
    title = ""
    statement_date = ""
    invested_value = None
    present_value = None
    unrealized_pnl = None
    has_pnl = False
    
    table_start_idx = -1
    
    # 1. Search for title & top summary block in the first 25 rows
    for i, row in enumerate(raw_rows[:25]):
        non_empty = [str(c).strip() for c in row if c is not None and str(c).strip() != ""]
        if not non_empty:
            continue
            
        row_str = " ".join(non_empty)
        row_str_lower = row_str.lower()
        
        # Check for title
        if not title and any(k in row_str_lower for k in ["holdings", "portfolio", "statement", "equity", "mutual fund", "demat", "stock", "cas", "cams", "kfintech", "wealth", "epf", "nps", "pension", "provident"]):
            title = non_empty[0] if len(non_empty) == 1 or not any(clean_num(x) for x in non_empty) else "Holdings Statement"
            date_match = re.search(r'(\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}|\d{1,2}\s+[A-Za-z]+\s+\d{4})', row_str)
            if date_match:
                statement_date = date_match.group(1)
                
        # Check for top summary key-values
        for j, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).strip().lower()
            
            if any(k in cell_str for k in ["invested value", "total cost", "total cost value", "total investment", "cost value", "invested amount"]):
                for next_cell in row[j+1:]:
                    num = clean_num(next_cell)
                    if num is not None and num > 0 and invested_value is None:
                        invested_value = num
                        has_pnl = True
                        break
            elif cell_str in ["invested", "investment"]:
                for next_cell in row[j+1:]:
                    num = clean_num(next_cell)
                    if num is not None and num > 0 and invested_value is None:
                        invested_value = num
                        has_pnl = True
                        break
                        
            if any(k in cell_str for k in ["present value", "current value", "market value", "total value", "portfolio value", "current market value", "total market value"]):
                for next_cell in row[j+1:]:
                    num = clean_num(next_cell)
                    if num is not None and num > 0 and present_value is None:
                        present_value = num
                        break
                        
            if any(k in cell_str for k in ["unrealized p&l", "unrealized pnl", "unrealised p&l", "unrealised pnl", "unrealized gain", "p&l", "pnl", "gain/loss", "total profit", "total returns"]):
                for next_cell in row[j+1:]:
                    num = clean_num(next_cell)
                    if num is not None and unrealized_pnl is None:
                        unrealized_pnl = num
                        has_pnl = True
                        break
        
        # Check if this row is the table header row
        header_keywords = [
            "sn", "sr no", "s.no", "s.n", "no", "description", "particulars", "item", "name", 
            "amount", "balance", "value", "total", "folio", "folio no", "isin", "scheme", 
            "scheme name", "cost value", "cost", "unit balance", "units", "nav date", "nav", 
            "nav (inr)", "market value", "market value (inr)", "registrar", "stock", "symbol", 
            "instrument", "company", "scrip", "qty", "quantity", "shares", "buy price", "avg price", 
            "ltp", "current price", "cur price", "market price", "invested", "current val", "market val", 
            "p&l", "profit"
        ]
        matches = sum(1 for kw in header_keywords if any(kw == str(c).strip().lower() or kw in str(c).strip().lower() for c in row if c is not None))
        
        if matches >= 2:
            table_start_idx = i
            break
            
    if not title:
        clean_name = filename.rsplit(".", 1)[0].replace("_", " ").replace("-", " ").title()
        title = clean_name
        
    headers = []
    data_rows = []
    
    total_row_cost = None
    total_row_market = None
    total_row_pnl = None
    
    if table_start_idx != -1:
        raw_headers = raw_rows[table_start_idx]
        last_col = 0
        for idx, h in enumerate(raw_headers):
            if h is not None and str(h).strip() != "":
                last_col = idx
                
        headers = [str(raw_headers[c]).strip() if c < len(raw_headers) and raw_headers[c] is not None and str(raw_headers[c]).strip() != "" else f"Column {c+1}" for c in range(last_col + 1)]
        
        # Identify Cost Value, Market Value, and General Amount columns
        cost_col = -1
        mkt_col = -1
        amt_col = -1
        pnl_col = -1
        
        for idx, h in enumerate(headers):
            h_lower = h.lower()
            if any(k in h_lower for k in ["cost value", "total cost", "buy value", "cost", "invested value", "invest amount", "invested"]) and cost_col == -1:
                cost_col = idx
                has_pnl = True
            elif any(k in h_lower for k in ["market value", "present value", "current value", "market val", "total val"]) and mkt_col == -1:
                mkt_col = idx
            elif any(k in h_lower for k in ["amount", "balance", "value"]) and amt_col == -1 and mkt_col == -1 and cost_col == -1:
                amt_col = idx
            elif any(k in h_lower for k in ["unrealized", "p&l", "pnl", "gain", "profit"]) and "%" not in h_lower and pnl_col == -1:
                pnl_col = idx
                has_pnl = True

        for row in raw_rows[table_start_idx + 1:]:
            if not any(row):
                continue
            row_cells = [str(row[c]).strip() if c < len(row) and row[c] is not None else "" for c in range(len(headers))]
            
            # Check if this row is a Total / Summary row
            if is_row_total_summary(row_cells):
                if cost_col != -1 and cost_col < len(row_cells):
                    val = clean_num(row_cells[cost_col])
                    if val and val > 0:
                        total_row_cost = val
                        has_pnl = True
                if mkt_col != -1 and mkt_col < len(row_cells):
                    val = clean_num(row_cells[mkt_col])
                    if val and val > 0:
                        total_row_market = val
                if pnl_col != -1 and pnl_col < len(row_cells):
                    val = clean_num(row_cells[pnl_col])
                    if val is not None:
                        total_row_pnl = val
                        has_pnl = True
                if amt_col != -1 and amt_col < len(row_cells) and total_row_market is None:
                    val = clean_num(row_cells[amt_col])
                    if val and val > 0:
                        total_row_market = val
                if total_row_cost is None or total_row_market is None:
                    nums_in_row = [clean_num(x) for x in row_cells if clean_num(x) is not None and clean_num(x) > 0]
                    if len(nums_in_row) >= 2:
                        if total_row_cost is None:
                            total_row_cost = nums_in_row[0]
                            has_pnl = True
                        if total_row_market is None: total_row_market = nums_in_row[1]
                continue  # EXCLUDE TOTAL ROW from data_rows!
                
            if any(row_cells):
                data_rows.append(row_cells)
    else:
        for i, row in enumerate(raw_rows):
            if any(row):
                headers = [str(c).strip() if c is not None else f"Column {idx+1}" for idx, c in enumerate(row)]
                for r in raw_rows[i+1:]:
                    if any(r):
                        row_c = [str(r[c]).strip() if c < len(r) and r[c] is not None else "" for c in range(len(headers))]
                        if not is_row_total_summary(row_c):
                            data_rows.append(row_c)
                break

    # If totals were explicitly present in the bottom Total row:
    if total_row_cost is not None and invested_value is None:
        invested_value = total_row_cost
        has_pnl = True
    if total_row_market is not None and present_value is None:
        present_value = total_row_market
    if total_row_pnl is not None and unrealized_pnl is None:
        unrealized_pnl = total_row_pnl
        has_pnl = True

    # If summary values were not found, sum from the clean data_rows (without total rows)
    if present_value is None or (invested_value is None and cost_col != -1):
        calc_invested = 0.0
        calc_present = 0.0
        calc_amt = 0.0
        calc_pnl = 0.0
        
        inv_col = -1
        cur_col = -1
        gen_col = -1
        pnl_col = -1
        
        for idx, h in enumerate(headers):
            h_lower = h.lower()
            if any(k in h_lower for k in ["cost value", "total cost", "buy value", "cost", "invested value", "invest amount", "invested"]) and inv_col == -1:
                inv_col = idx
                has_pnl = True
            elif any(k in h_lower for k in ["market value", "present value", "current value", "market val", "cur val", "total val"]) and cur_col == -1:
                cur_col = idx
            elif any(k in h_lower for k in ["amount", "balance", "value"]) and gen_col == -1 and cur_col == -1:
                gen_col = idx
            elif any(k in h_lower for k in ["unrealized", "p&l", "pnl", "gain", "profit"]) and "%" not in h_lower and pnl_col == -1:
                pnl_col = idx
                has_pnl = True
                
        for row in data_rows:
            if inv_col != -1 and inv_col < len(row):
                v = clean_num(row[inv_col])
                if v: calc_invested += v
            if cur_col != -1 and cur_col < len(row):
                v = clean_num(row[cur_col])
                if v: calc_present += v
            elif gen_col != -1 and gen_col < len(row):
                v = clean_num(row[gen_col])
                if v: calc_amt += v
            if pnl_col != -1 and pnl_col < len(row):
                v = clean_num(row[pnl_col])
                if v: calc_pnl += v
                
        if inv_col != -1 and invested_value is None and calc_invested > 0:
            invested_value = calc_invested
        if cur_col != -1 and present_value is None and calc_present > 0:
            present_value = calc_present
        elif gen_col != -1 and present_value is None and calc_amt > 0:
            present_value = calc_amt
            if invested_value is None:
                invested_value = calc_amt
                has_pnl = False
                
        if unrealized_pnl is None and calc_pnl != 0:
            unrealized_pnl = calc_pnl
            has_pnl = True

    if present_value is None:
        present_value = 0.0
    if invested_value is None:
        invested_value = present_value
    if unrealized_pnl is None:
        if has_pnl:
            unrealized_pnl = present_value - invested_value
        else:
            unrealized_pnl = 0.0
        
    pnl_percentage = round((unrealized_pnl / invested_value * 100), 2) if (has_pnl and invested_value > 0) else 0.0
    
    return {
        "title": title,
        "statement_date": statement_date,
        "invested_value": round(invested_value),
        "present_value": round(present_value),
        "unrealized_pnl": round(unrealized_pnl),
        "pnl_percentage": pnl_percentage,
        "has_pnl": 1 if has_pnl else 0,
        "headers": headers,
        "data_rows": data_rows,
        "row_count": len(data_rows)
    }


def parse_holdings_file(file_obj, filename):
    lower_filename = filename.lower()
    raw_rows = []
    
    if lower_filename.endswith(".csv"):
        try:
            content = file_obj.read().decode("utf-8-sig", errors="replace")
            reader = csv.reader(io.StringIO(content))
            raw_rows = list(reader)
        except Exception:
            return None
    elif lower_filename.endswith((".xlsx", ".xls")):
        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            ws = wb.active
            raw_rows = list(ws.iter_rows(values_only=True))
        except Exception:
            return None
    else:
        return None
        
    return parse_holdings_data(raw_rows, filename)


# ============================================================
# DASHBOARD
# ============================================================

@app.route("/")
@login_required
def home():
    conn = get_db()
    user_id = session["user_id"]

    transactions = conn.execute("""
        SELECT *
        FROM transactions
        WHERE user_id = ?
        ORDER BY date DESC, id DESC
    """, (user_id,)).fetchall()

    # 1. Income
    total_income = conn.execute("""
        SELECT COALESCE(SUM(amount), 0)
        FROM transactions
        WHERE user_id = ?
        AND type = 'Income'
    """, (user_id,)).fetchone()[0]

    # 2. Expense
    total_expenses = conn.execute("""
        SELECT COALESCE(SUM(amount), 0)
        FROM transactions
        WHERE user_id = ?
        AND type = 'Expense'
        AND LOWER(TRIM(category)) NOT IN (
            'investment', 'investments', 'sip', 'mutual fund', 'mutual funds',
            'stocks', 'equity', 'bonds', 'crypto', 'gold', 'insurance',
            'fixed deposit', 'fd', 'ppf', 'nps', 'reit'
        )
    """, (user_id,)).fetchone()[0]

    # 3. Investment
    total_investments = conn.execute("""
        SELECT COALESCE(SUM(amount), 0)
        FROM transactions
        WHERE user_id = ?
        AND (
            type = 'Investment'
            OR (
                type = 'Expense'
                AND LOWER(TRIM(category)) IN (
                    'investment', 'investments', 'sip', 'mutual fund', 'mutual funds',
                    'stocks', 'equity', 'bonds', 'crypto', 'gold', 'insurance',
                    'fixed deposit', 'fd', 'ppf', 'nps', 'reit'
                )
            )
        )
    """, (user_id,)).fetchone()[0]

    # 4. Debit = Expense + Investment
    total_debit = total_expenses + total_investments

    # 5. Balance = Income - Debit
    balance = total_income - total_debit

    categories = [row["category"] for row in conn.execute("""
        SELECT DISTINCT category FROM transactions
        WHERE user_id = ? AND category IS NOT NULL AND TRIM(category) != ''
        ORDER BY category ASC
    """, (user_id,)).fetchall()]

    accounts = [row["account"] for row in conn.execute("""
        SELECT DISTINCT account FROM transactions
        WHERE user_id = ? AND account IS NOT NULL AND TRIM(account) != ''
        ORDER BY account ASC
    """, (user_id,)).fetchall()]

    conn.close()

    today_str = datetime.now().strftime("%Y-%m-%d")

    return render_template(
        "index.html",
        transactions=transactions,
        total_income=float(total_income or 0),
        total_debit=float(total_debit or 0),
        total_expenses=float(total_expenses or 0),
        total_investments=float(total_investments or 0),
        balance=float(balance or 0),
        categories=categories,
        accounts=accounts,
        today=today_str
    )


# ============================================================
# HOLDINGS & MULTI-FILE PORTFOLIO TAB
# ============================================================

@app.route("/holdings")
@login_required
def holdings():
    conn = get_db()
    user_id = session["user_id"]

    statements = conn.execute("""
        SELECT id, filename, title, upload_date, statement_date, invested_value, present_value, unrealized_pnl, pnl_percentage, has_pnl, row_count
        FROM holdings_statements
        WHERE user_id = ?
        ORDER BY id DESC
    """, (user_id,)).fetchall()

    total_present_value = sum(s["present_value"] for s in statements)
    total_invested_value = sum(s["invested_value"] for s in statements)
    total_unrealized_pnl = sum(s["unrealized_pnl"] for s in statements if s["has_pnl"])
    
    pnl_invested_total = sum(s["invested_value"] for s in statements if s["has_pnl"])
    overall_pnl_percentage = round((total_unrealized_pnl / pnl_invested_total * 100), 2) if pnl_invested_total > 0 else 0.0

    conn.close()

    return render_template(
        "holdings.html",
        statements=statements,
        total_present_value=total_present_value,
        total_invested_value=total_invested_value,
        total_unrealized_pnl=total_unrealized_pnl,
        overall_pnl_percentage=overall_pnl_percentage
    )


@app.route("/holdings/upload", methods=["POST"])
@login_required
def upload_holdings():
    files = request.files.getlist("files")
    if not files or all(f.filename == "" for f in files):
        flash("Please select at least one file to import.", "warning")
        return redirect(url_for("holdings"))

    conn = get_db()
    user_id = session["user_id"]
    success_count = 0
    now_str = datetime.now().strftime("%d %b %Y, %I:%M %p")

    for f in files:
        if not f or f.filename == "":
            continue
        parsed = parse_holdings_file(f, f.filename)
        if not parsed:
            continue

        conn.execute("""
            INSERT INTO holdings_statements (
                user_id, filename, title, upload_date, statement_date,
                invested_value, present_value, unrealized_pnl, pnl_percentage, has_pnl,
                headers_json, data_json, row_count
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            user_id,
            f.filename,
            parsed["title"],
            now_str,
            parsed["statement_date"],
            parsed["invested_value"],
            parsed["present_value"],
            parsed["unrealized_pnl"],
            parsed["pnl_percentage"],
            parsed["has_pnl"],
            json.dumps(parsed["headers"]),
            json.dumps(parsed["data_rows"]),
            parsed["row_count"]
        ))
        success_count += 1

    conn.commit()
    conn.close()

    if success_count > 0:
        flash(f"Successfully imported {success_count} statement{'s' if success_count > 1 else ''}!", "success")
    else:
        flash("Could not parse the selected file(s). Please make sure they are valid Excel or CSV files.", "danger")

    return redirect(url_for("holdings"))


@app.route("/holdings/<int:statement_id>")
@login_required
def holdings_detail(statement_id):
    conn = get_db()
    user_id = session["user_id"]

    stmt = conn.execute("""
        SELECT * FROM holdings_statements
        WHERE id = ? AND user_id = ?
    """, (statement_id, user_id)).fetchone()

    conn.close()

    if not stmt:
        flash("Statement not found.", "danger")
        return redirect(url_for("holdings"))

    headers = json.loads(stmt["headers_json"] or "[]")
    data_rows = json.loads(stmt["data_json"] or "[]")

    return render_template(
        "holdings_detail.html",
        statement=stmt,
        headers=headers,
        data_rows=data_rows
    )


@app.route("/holdings/delete/<int:statement_id>", methods=["POST"])
@login_required
def delete_holdings_statement(statement_id):
    conn = get_db()
    conn.execute("DELETE FROM holdings_statements WHERE id = ? AND user_id = ?", (statement_id, session["user_id"]))
    conn.commit()
    conn.close()

    flash("Statement deleted.", "info")
    return redirect(url_for("holdings"))


@app.route("/holdings/delete-all", methods=["POST"])
@login_required
def delete_all_holdings():
    conn = get_db()
    conn.execute("DELETE FROM holdings_statements WHERE user_id = ?", (session["user_id"],))
    conn.commit()
    conn.close()

    flash("All holdings statements have been cleared.", "warning")
    return redirect(url_for("holdings"))


# ============================================================
# ADD TRANSACTION
# ============================================================

@app.route("/add", methods=["POST"])
@login_required
def add_transaction():
    date_val = normalize_date(request.form.get("date", "").strip())
    category = request.form.get("category", "Other").strip() or "Other"
    subcategory = request.form.get("subcategory", "").strip()
    amount = clean_amount(request.form.get("amount", "0"))
    description = request.form.get("description", "").strip()
    place = request.form.get("place", "").strip()
    account = request.form.get("account", "").strip()
    remarks = request.form.get("remarks", "").strip()

    raw_type = request.form.get("type", "Expense")
    transaction_type = get_real_type(raw_type, category)

    if amount <= 0:
        flash("Amount must be greater than zero.", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    conn.execute("""
        INSERT INTO transactions (
            user_id, date, type, category, subcategory, amount, description, place, account, remarks
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        session["user_id"],
        date_val,
        transaction_type,
        category,
        subcategory,
        amount,
        description,
        place,
        account,
        remarks
    ))
    conn.commit()
    conn.close()

    flash("Transaction added successfully!", "success")
    return redirect(url_for("home"))


# ============================================================
# EDIT TRANSACTION
# ============================================================

@app.route("/edit/<int:transaction_id>", methods=["POST"])
@login_required
def edit_transaction(transaction_id):
    conn = get_db()
    tx = conn.execute("SELECT * FROM transactions WHERE id = ? AND user_id = ?", (transaction_id, session["user_id"])).fetchone()
    
    if not tx:
        conn.close()
        flash("Transaction not found or access denied.", "danger")
        return redirect(url_for("home"))

    date_val = normalize_date(request.form.get("date", "").strip())
    category = request.form.get("category", "Other").strip() or "Other"
    subcategory = request.form.get("subcategory", "").strip()
    amount = clean_amount(request.form.get("amount", "0"))
    description = request.form.get("description", "").strip()
    place = request.form.get("place", "").strip()
    account = request.form.get("account", "").strip()
    remarks = request.form.get("remarks", "").strip()

    raw_type = request.form.get("type", "Expense")
    transaction_type = get_real_type(raw_type, category)

    if amount <= 0:
        conn.close()
        flash("Amount must be greater than zero.", "danger")
        return redirect(url_for("home"))

    conn.execute("""
        UPDATE transactions
        SET date = ?, type = ?, category = ?, subcategory = ?, amount = ?, description = ?, place = ?, account = ?, remarks = ?
        WHERE id = ? AND user_id = ?
    """, (
        date_val,
        transaction_type,
        category,
        subcategory,
        amount,
        description,
        place,
        account,
        remarks,
        transaction_id,
        session["user_id"]
    ))
    conn.commit()
    conn.close()

    flash("Transaction updated successfully!", "success")
    return redirect(url_for("home"))


# ============================================================
# DELETE ONE
# ============================================================

@app.route("/delete/<int:transaction_id>", methods=["POST"])
@login_required
def delete_transaction(transaction_id):
    conn = get_db()
    conn.execute("""
        DELETE FROM transactions
        WHERE id = ? AND user_id = ?
    """, (transaction_id, session["user_id"]))
    conn.commit()
    conn.close()

    flash("Transaction deleted.", "info")
    return redirect(url_for("home"))


# ============================================================
# DELETE ALL
# ============================================================

@app.route("/delete-all", methods=["POST"])
@login_required
def delete_all_transactions():
    conn = get_db()
    conn.execute("""
        DELETE FROM transactions
        WHERE user_id = ?
    """, (session["user_id"],))
    conn.commit()
    conn.close()

    flash("All transactions have been deleted.", "warning")
    return redirect(url_for("home"))


# ============================================================
# EXPORT CSV
# ============================================================

@app.route("/export/csv")
@login_required
def export_csv():
    conn = get_db()
    transactions = conn.execute("""
        SELECT date, type, category, subcategory, amount, description, place, account, remarks
        FROM transactions
        WHERE user_id = ?
        ORDER BY date DESC, id DESC
    """, (session["user_id"],)).fetchall()
    conn.close()

    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output)

    writer.writerow(["Tran Date", "Type", "Category", "Sub Category", "Amount", "Description", "Place", "Account", "Remarks"])

    for tx in transactions:
        writer.writerow([
            tx["date"],
            tx["type"],
            tx["category"],
            tx["subcategory"] or "",
            tx["amount"],
            tx["description"] or "",
            tx["place"] or "",
            tx["account"] or "",
            tx["remarks"] or ""
        ])

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename=ExpenseTracker_Export_{datetime.now().strftime('%Y%m%d')}.csv"}
    )


# ============================================================
# EXPORT EXCEL
# ============================================================

@app.route("/export/excel")
@login_required
def export_excel():
    conn = get_db()
    transactions = conn.execute("""
        SELECT date, type, category, subcategory, amount, description, place, account, remarks
        FROM transactions
        WHERE user_id = ?
        ORDER BY date DESC, id DESC
    """, (session["user_id"],)).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["Tran Date", "Type", "Category", "Sub Category", "Amount", "Description", "Place", "Account", "Remarks"]
    ws.append(headers)

    header_fill = PatternFill(start_color="2F7A6F", end_color="2F7A6F", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='E4EAE7'),
        right=Side(style='thin', color='E4EAE7'),
        top=Side(style='thin', color='E4EAE7'),
        bottom=Side(style='thin', color='E4EAE7')
    )

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, tx in enumerate(transactions, start=2):
        ws.append([
            tx["date"],
            tx["type"],
            tx["category"],
            tx["subcategory"] or "",
            tx["amount"],
            tx["description"] or "",
            tx["place"] or "",
            tx["account"] or "",
            tx["remarks"] or ""
        ])
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx == 5:
                cell.number_format = '₹#,##0'

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename=ExpenseTracker_Export_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )


# ============================================================
# ANALYTICS
# ============================================================

@app.route("/analytics")
@login_required
def analytics():
    conn = get_db()
    user_id = session["user_id"]

    selected_type = request.args.get("type", "Debit")
    if selected_type not in ["Income", "Debit", "Expense", "Investment"]:
        selected_type = "Debit"

    selected_category = request.args.get("category", "").strip()
    timeframe = request.args.get("timeframe", "all")

    date_filter = ""
    date_params = []
    now = datetime.now()

    if timeframe == "month":
        month_start = now.strftime("%Y-%m-01")
        date_filter = "AND date >= ?"
        date_params.append(month_start)
    elif timeframe == "last_month":
        first_of_this_month = now.replace(day=1)
        last_month_end = first_of_this_month - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)
        date_filter = "AND date >= ? AND date <= ?"
        date_params.extend([last_month_start.strftime("%Y-%m-%d"), last_month_end.strftime("%Y-%m-%d")])
    elif timeframe == "year":
        year_start = now.strftime("%Y-01-01")
        date_filter = "AND date >= ?"
        date_params.append(year_start)

    # 1. Total Income
    total_income = conn.execute(f"""
        SELECT COALESCE(SUM(amount), 0)
        FROM transactions
        WHERE user_id = ?
        AND type = 'Income'
        {date_filter}
    """, [user_id] + date_params).fetchone()[0]

    # 2. Total Expenses
    total_expenses = conn.execute(f"""
        SELECT COALESCE(SUM(amount), 0)
        FROM transactions
        WHERE user_id = ?
        AND type = 'Expense'
        AND LOWER(TRIM(category)) NOT IN (
            'investment', 'investments', 'sip', 'mutual fund', 'mutual funds',
            'stocks', 'equity', 'bonds', 'crypto', 'gold', 'insurance',
            'fixed deposit', 'fd', 'ppf', 'nps', 'reit'
        )
        {date_filter}
    """, [user_id] + date_params).fetchone()[0]

    # 3. Total Investments
    total_investments = conn.execute(f"""
        SELECT COALESCE(SUM(amount), 0)
        FROM transactions
        WHERE user_id = ?
        AND (
            type = 'Investment'
            OR (
                type = 'Expense'
                AND LOWER(TRIM(category)) IN (
                    'investment', 'investments', 'sip', 'mutual fund', 'mutual funds',
                    'stocks', 'equity', 'bonds', 'crypto', 'gold', 'insurance',
                    'fixed deposit', 'fd', 'ppf', 'nps', 'reit'
                )
            )
        )
        {date_filter}
    """, [user_id] + date_params).fetchone()[0]

    # 4. Total Debit = Expense + Investment
    total_debit = total_expenses + total_investments

    # 5. Balance = Income - Debit
    balance = total_income - total_debit

    category_data = []
    subcategory_data = []

    # --------------------------------------------------------
    # 1. DEBIT (Expense + Investment)
    # --------------------------------------------------------
    if selected_type == "Debit":
        debit_txs = conn.execute(f"""
            SELECT * FROM transactions
            WHERE user_id = ?
            AND (type = 'Expense' OR type = 'Investment' OR type != 'Income')
            {date_filter}
            ORDER BY amount DESC
        """, [user_id] + date_params).fetchall()

        if not selected_category:
            cat_map = {}
            for r in debit_txs:
                cat_val = r["category"] or "Other"
                if is_investment_category(cat_val) or r["type"] == "Investment":
                    c_name, _ = get_investment_category_and_sub(r)
                else:
                    c_name = cat_val.strip() or "Other"
                cat_map[c_name] = cat_map.get(c_name, 0.0) + float(r["amount"] or 0)

            category_data = [{"category": k, "total": round(v, 2)} for k, v in sorted(cat_map.items(), key=lambda x: x[1], reverse=True)]
            selected_type_total = sum(item["total"] for item in category_data) or 1.0
            for item in category_data:
                item["percentage"] = round((item["total"] / selected_type_total) * 100, 1)
        else:
            sub_map = {}
            for r in debit_txs:
                cat_val = r["category"] or "Other"
                if is_investment_category(cat_val) or r["type"] == "Investment":
                    c_name, s_name = get_investment_category_and_sub(r)
                else:
                    c_name = cat_val.strip() or "Other"
                    s_name = (r["subcategory"] or r["description"] or "General").strip() or "General"

                if c_name.lower() == selected_category.lower() or selected_category.lower() in [c_name.lower(), str(r["category"]).lower(), str(r["subcategory"]).lower()]:
                    sub_map[s_name] = sub_map.get(s_name, 0.0) + float(r["amount"] or 0)

            subcategory_data = [{"subcategory": k, "total": round(v, 2)} for k, v in sorted(sub_map.items(), key=lambda x: x[1], reverse=True)]
            sub_total = sum(item["total"] for item in subcategory_data) or 1.0
            for item in subcategory_data:
                item["percentage"] = round((item["total"] / sub_total) * 100, 1)

    # --------------------------------------------------------
    # 2. INVESTMENT HIERARCHY
    # --------------------------------------------------------
    elif selected_type == "Investment":
        inv_txs = conn.execute(f"""
            SELECT * FROM transactions
            WHERE user_id = ?
            AND (
                type = 'Investment'
                OR (
                    type = 'Expense'
                    AND LOWER(TRIM(category)) IN (
                        'investment', 'investments', 'sip', 'mutual fund', 'mutual funds',
                        'stocks', 'equity', 'bonds', 'crypto', 'gold', 'insurance',
                        'fixed deposit', 'fd', 'ppf', 'nps', 'reit'
                    )
                )
            )
            {date_filter}
            ORDER BY amount DESC
        """, [user_id] + date_params).fetchall()

        if not selected_category:
            cat_map = {}
            for r in inv_txs:
                c, s = get_investment_category_and_sub(r)
                cat_map[c] = cat_map.get(c, 0.0) + float(r["amount"] or 0)
            
            category_data = [{"category": k, "total": round(v, 2)} for k, v in sorted(cat_map.items(), key=lambda x: x[1], reverse=True)]
            selected_type_total = sum(item["total"] for item in category_data) or 1.0
            for item in category_data:
                item["percentage"] = round((item["total"] / selected_type_total) * 100, 1)
        else:
            sub_map = {}
            for r in inv_txs:
                c, s = get_investment_category_and_sub(r)
                if c.lower() == selected_category.lower() or selected_category.lower() in [c.lower(), str(r["category"]).lower(), str(r["subcategory"]).lower()]:
                    sub_map[s] = sub_map.get(s, 0.0) + float(r["amount"] or 0)
            
            subcategory_data = [{"subcategory": k, "total": round(v, 2)} for k, v in sorted(sub_map.items(), key=lambda x: x[1], reverse=True)]
            sub_total = sum(item["total"] for item in subcategory_data) or 1.0
            for item in subcategory_data:
                item["percentage"] = round((item["total"] / sub_total) * 100, 1)

    # --------------------------------------------------------
    # 3. EXPENSE HIERARCHY
    # --------------------------------------------------------
    elif selected_type == "Expense":
        if not selected_category:
            category_rows = conn.execute(f"""
                SELECT category, SUM(amount) AS total
                FROM transactions
                WHERE user_id = ?
                AND type = 'Expense'
                AND LOWER(TRIM(category)) NOT IN (
                    'investment', 'investments', 'sip', 'mutual fund', 'mutual funds',
                    'stocks', 'equity', 'bonds', 'crypto', 'gold', 'insurance',
                    'fixed deposit', 'fd', 'ppf', 'nps', 'reit'
                )
                {date_filter}
                GROUP BY category
                ORDER BY total DESC
            """, [user_id] + date_params).fetchall()

            category_data = [{"category": row["category"] or "Other", "total": round(float(row["total"] or 0), 2)} for row in category_rows]
            selected_type_total = sum(item["total"] for item in category_data) or 1.0
            for item in category_data:
                item["percentage"] = round((item["total"] / selected_type_total) * 100, 1)
        else:
            sub_rows = conn.execute(f"""
                SELECT COALESCE(NULLIF(TRIM(subcategory), ''), NULLIF(TRIM(description), ''), 'General') AS subcategory, SUM(amount) AS total
                FROM transactions
                WHERE user_id = ?
                AND type = 'Expense'
                AND LOWER(TRIM(category)) NOT IN (
                    'investment', 'investments', 'sip', 'mutual fund', 'mutual funds',
                    'stocks', 'equity', 'bonds', 'crypto', 'gold', 'insurance',
                    'fixed deposit', 'fd', 'ppf', 'nps', 'reit'
                )
                AND category = ?
                {date_filter}
                GROUP BY COALESCE(NULLIF(TRIM(subcategory), ''), NULLIF(TRIM(description), ''), 'General')
                ORDER BY total DESC
            """, [user_id, selected_category] + date_params).fetchall()

            subcategory_data = [{
                "subcategory": r["subcategory"],
                "total": round(float(r["total"] or 0), 2)
            } for r in sub_rows]
            sub_total = sum(item["total"] for item in subcategory_data) or 1.0
            for item in subcategory_data:
                item["percentage"] = round((item["total"] / sub_total) * 100, 1)

    # --------------------------------------------------------
    # 4. INCOME HIERARCHY
    # --------------------------------------------------------
    else:
        if not selected_category:
            category_rows = conn.execute(f"""
                SELECT category, SUM(amount) AS total
                FROM transactions
                WHERE user_id = ?
                AND type = 'Income'
                {date_filter}
                GROUP BY category
                ORDER BY total DESC
            """, [user_id] + date_params).fetchall()

            category_data = [{"category": row["category"] or "Other", "total": round(float(row["total"] or 0), 2)} for row in category_rows]
            selected_type_total = sum(item["total"] for item in category_data) or 1.0
            for item in category_data:
                item["percentage"] = round((item["total"] / selected_type_total) * 100, 1)
        else:
            sub_rows = conn.execute(f"""
                SELECT COALESCE(NULLIF(TRIM(subcategory), ''), NULLIF(TRIM(description), ''), 'General') AS subcategory, SUM(amount) AS total
                FROM transactions
                WHERE user_id = ?
                AND type = 'Income'
                AND category = ?
                {date_filter}
                GROUP BY COALESCE(NULLIF(TRIM(subcategory), ''), NULLIF(TRIM(description), ''), 'General')
                ORDER BY total DESC
            """, [user_id, selected_category] + date_params).fetchall()

            subcategory_data = [{
                "subcategory": r["subcategory"],
                "total": round(float(r["total"] or 0), 2)
            } for r in sub_rows]
            sub_total = sum(item["total"] for item in subcategory_data) or 1.0
            for item in subcategory_data:
                item["percentage"] = round((item["total"] / sub_total) * 100, 1)

    conn.close()

    return render_template(
        "analytics.html",
        selected_type=selected_type,
        selected_category=selected_category,
        timeframe=timeframe,
        category_data=category_data,
        subcategory_data=subcategory_data,
        total_income=float(total_income or 0),
        total_debit=float(total_debit or 0),
        total_expenses=float(total_expenses or 0),
        total_investments=float(total_investments or 0),
        balance=float(balance or 0)
    )


# ============================================================
# IMPORT CSV / EXCEL (TRANSACTIONS)
# ============================================================

@app.route("/import", methods=["GET", "POST"])
@login_required
def import_excel():
    if request.method == "GET":
        return render_template("import.html")

    if "file" not in request.files:
        flash("Please select a file to import.", "danger")
        return render_template("import.html")

    file = request.files["file"]
    if file.filename == "":
        flash("Please select a file.", "danger")
        return render_template("import.html")

    filename = file.filename
    lower_filename = filename.lower()
    rows = []

    if lower_filename.endswith(".csv"):
        try:
            content = file.read().decode("utf-8-sig", errors="replace")
            reader = csv.reader(io.StringIO(content))
            raw_rows = list(reader)
            if raw_rows:
                headers = [h.strip() for h in raw_rows[0]]
                for values_row in raw_rows[1:]:
                    if not any(values_row):
                        continue
                    row = {}
                    for idx, header in enumerate(headers):
                        if idx < len(values_row):
                            h_name = header
                            if h_name in row:
                                h_name = f"{header}_{idx}"
                            row[h_name] = values_row[idx]
                    rows.append(row)
        except Exception as error:
            flash(f"Could not read CSV file: {str(error)}", "danger")
            return render_template("import.html")

    elif lower_filename.endswith((".xlsx", ".xls")):
        try:
            workbook = openpyxl.load_workbook(file, data_only=True)
            sheet = workbook.active
            values = list(sheet.values)
            if not values:
                flash("The Excel file is empty.", "warning")
                return render_template("import.html")

            headers = [str(h).strip() if h is not None else f"Col_{i}" for i, h in enumerate(values[0])]
            for values_row in values[1:]:
                if not any(values_row):
                    continue
                row = {}
                for idx, header in enumerate(headers):
                    if idx < len(values_row):
                        h_name = header
                        if h_name in row:
                            h_name = f"{header}_{idx}"
                        row[h_name] = values_row[idx]
                rows.append(row)
        except Exception as error:
            flash(f"Could not read Excel file: {str(error)}", "danger")
            return render_template("import.html")
    else:
        flash("Please upload a valid CSV or Excel (.xlsx, .xls) file.", "danger")
        return render_template("import.html")

    if not rows:
        flash("The uploaded file contains no transaction records.", "warning")
        return render_template("import.html")

    conn = get_db()
    imported_count = 0

    for row in rows:
        date_raw = (
            row.get("Tran Date") or row.get("Date") or row.get("Transaction Date")
            or row.get("TRAN DATE") or row.get("date") or ""
        )
        if not date_raw:
            continue

        date_val = normalize_date(date_raw)

        amount_raw = (
            row.get("Amount") or row.get("AMOUNT") or row.get("amount")
            or row.get("Debit") or row.get("Credit") or 0
        )
        amount = clean_amount(amount_raw)
        if amount <= 0:
            continue

        category = str(
            row.get("Category") or row.get("CATEGORY") or row.get("category") or "Other"
        ).strip() or "Other"

        raw_type = (
            row.get("Type") or row.get("TYPE") or row.get("type") or "Expense"
        )
        transaction_type = get_real_type(raw_type, category)

        subcategory = str(
            row.get("Sub Category") or row.get("Subcategory") or row.get("SUB CATEGORY") or row.get("subcategory") or ""
        ).strip()

        desc_keys = [k for k in row.keys() if k.lower().startswith("description")]
        description = ""
        remarks_fallback = ""
        if len(desc_keys) == 1:
            description = str(row.get(desc_keys[0]) or "").strip()
        elif len(desc_keys) >= 2:
            description = str(row.get(desc_keys[0]) or "").strip()
            remarks_fallback = str(row.get(desc_keys[1]) or "").strip()

        place = str(row.get("Place") or row.get("PLACE") or row.get("Location") or "").strip()
        account = str(row.get("Account") or row.get("ACCOUNT") or row.get("Bank") or "").strip()
        remarks = str(row.get("Remarks") or row.get("REMARKS") or row.get("Notes") or remarks_fallback or "").strip()

        if not subcategory and remarks:
            subcategory = remarks

        conn.execute("""
            INSERT INTO transactions (
                user_id, date, type, category, subcategory, amount, description, place, account, remarks
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            session["user_id"],
            date_val,
            transaction_type,
            category,
            subcategory,
            amount,
            description,
            place,
            account,
            remarks
        ))
        imported_count += 1

    conn.commit()
    conn.close()

    flash(f"Successfully imported {imported_count} transactions from {filename}!", "success")
    return render_template(
        "import_success.html",
        filename=filename,
        imported_count=imported_count
    )


# ============================================================
# SIGNUP / AUTH
# ============================================================

@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "GET":
        if "user_id" in session:
            return redirect(url_for("home"))
        return render_template("signup.html")

    name = request.form.get("name", "").strip()
    email = request.form.get("email", "").strip().lower()
    password = request.form.get("password", "")

    if not name or not email or not password:
        flash("Please fill in all fields.", "danger")
        return render_template("signup.html", name=name, email=email)

    if len(password) < 8:
        flash("Password must be at least 8 characters.", "danger")
        return render_template("signup.html", name=name, email=email)

    conn = get_db()
    existing = conn.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()
    if existing:
        conn.close()
        flash("An account with this email already exists.", "danger")
        return render_template("signup.html", name=name, email=email)

    password_hash = generate_password_hash(password)
    cursor = conn.execute("INSERT INTO users (name, email, password) VALUES (?, ?, ?)", (name, email, password_hash))
    user_id = cursor.lastrowid
    conn.commit()
    conn.close()

    session.clear()
    session["user_id"] = user_id
    session["name"] = name
    session["email"] = email

    flash(f"Welcome to Expense Tracker, {name}! 🎉", "success")
    return redirect(url_for("home"))


# ============================================================
# LOGIN
# ============================================================

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        if "user_id" in session:
            return redirect(url_for("home"))
        return render_template("login.html")

    email = request.form.get("email", "").strip().lower()
    password = request.form.get("password", "")

    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
    conn.close()

    if user is None or not check_password_hash(user["password"], password):
        flash("Invalid email or password.", "danger")
        return render_template("login.html", email=email)

    session.clear()
    session["user_id"] = user["id"]
    session["name"] = user["name"]
    session["email"] = user["email"]

    flash(f"Welcome back, {user['name']}!", "success")
    return redirect(url_for("home"))


# ============================================================
# LOGOUT
# ============================================================

@app.route("/logout")
def logout():
    session.clear()
    flash("You have been signed out.", "info")
    return redirect(url_for("login"))


# ============================================================
# RUN
# ============================================================

if __name__ == "__main__":
    init_db()
    app.run(debug=True, port=5000)
